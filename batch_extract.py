#!/usr/bin/env python3
"""
Batch extract multiple bank statements into a single Excel file.
Creates:
1. All Transactions sheet - consolidated view of all transactions
2. Summary sheet - extraction statistics
3. Individual statement sheets (optional)
"""
import sys
import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from src.pipeline import ExtractionPipeline
from src.utils import setup_logger

logger = setup_logger()


def sanitize_sheet_name(name: str) -> str:
    """
    Excel sheet names have restrictions:
    - Max 31 characters
    - Cannot contain: \ / ? * [ ]
    """
    # Remove file extension
    name = Path(name).stem

    # Replace invalid characters
    for char in ['\\', '/', '?', '*', '[', ']']:
        name = name.replace(char, '_')

    # Truncate to 31 chars
    if len(name) > 31:
        name = name[:31]

    return name


def batch_extract(folder_path: str, output_path: str):
    """
    Extract all PDF statements in a folder and combine into one Excel file.

    Args:
        folder_path: Path to folder containing statement PDFs
        output_path: Path to output Excel file
    """
    folder = Path(folder_path)
    output = Path(output_path)

    if not folder.exists():
        logger.error(f"Folder not found: {folder}")
        return False

    # Find all PDF files
    pdf_files = sorted(folder.glob("*.pdf"))

    if not pdf_files:
        logger.error(f"No PDF files found in {folder}")
        return False

    logger.info(f"Found {len(pdf_files)} PDF files to process")
    logger.info("=" * 80)

    # Initialize pipeline
    pipeline = ExtractionPipeline()

    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Track results and all transactions
    results = []
    all_transactions = []  # For consolidated sheet

    # Process each file
    for idx, pdf_file in enumerate(pdf_files, 1):
        logger.info(f"\n[{idx}/{len(pdf_files)}] Processing: {pdf_file.name}")

        try:
            # Extract
            result = pipeline.process(pdf_file)

            if not result.success or not result.transactions:
                logger.warning(f"  ⚠ Failed or no transactions: {pdf_file.name}")
                results.append({
                    'file': pdf_file.name,
                    'success': False,
                    'transactions': 0,
                    'error': result.error_message or 'No transactions found'
                })
                continue

            # Create sheet name from filename
            sheet_name = sanitize_sheet_name(pdf_file.name)

            # Ensure unique sheet name
            base_name = sheet_name
            counter = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base_name[:28]}_{counter}"
                counter += 1

            # Create worksheet
            ws = wb.create_sheet(title=sheet_name)

            # Convert transactions to DataFrame and add to consolidated list
            txn_data = []
            for txn in result.transactions:
                txn_row = {
                    'Statement': pdf_file.stem,  # Add statement identifier
                    'Date': txn.date.strftime('%Y-%m-%d'),
                    'Description': txn.description,
                    'Money In': txn.money_in if txn.money_in else None,
                    'Money Out': txn.money_out if txn.money_out else None,
                    'Balance': txn.balance,
                    'Confidence %': txn.confidence
                }
                txn_data.append(txn_row)
                all_transactions.append(txn_row)  # Add to consolidated list

            df = pd.DataFrame(txn_data)

            # Add metadata header (first 3 rows)
            bank_name = result.statement.bank_name if result.statement else "Unknown"
            account = result.statement.account_number if result.statement else "N/A"
            period = ""
            if result.statement and result.statement.statement_start_date and result.statement.statement_end_date:
                period = f"{result.statement.statement_start_date.strftime('%Y-%m-%d')} to {result.statement.statement_end_date.strftime('%Y-%m-%d')}"

            ws['A1'] = f"Bank: {bank_name}"
            ws['D1'] = f"Account: {account}"
            ws['A2'] = f"Period: {period}"
            ws['D2'] = f"Transactions: {len(result.transactions)}"
            ws['A3'] = f"Reconciled: {'✓ Yes' if result.balance_reconciled else '✗ No'}"
            ws['D3'] = f"Confidence: {result.confidence_score:.1f}%"

            # Style header
            for cell in ['A1', 'D1', 'A2', 'D2', 'A3', 'D3']:
                ws[cell].font = Font(bold=True, size=11)

            # Add column headers at row 5 (individual sheet doesn't need Statement column)
            headers = ['Date', 'Description', 'Money In', 'Money Out', 'Balance', 'Confidence %']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_idx, value=header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            # Add data starting at row 6
            for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 6):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Set column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12

            logger.info(f"  ✓ Success: {len(result.transactions)} transactions, {result.confidence_score:.1f}% confidence")

            results.append({
                'file': pdf_file.name,
                'success': True,
                'transactions': len(result.transactions),
                'confidence': result.confidence_score,
                'reconciled': result.balance_reconciled
            })

        except Exception as e:
            logger.error(f"  ✗ Error processing {pdf_file.name}: {str(e)}")
            results.append({
                'file': pdf_file.name,
                'success': False,
                'transactions': 0,
                'error': str(e)
            })

    # Create consolidated "All Transactions" sheet
    logger.info("\n" + "=" * 80)
    logger.info("Creating consolidated transactions sheet...")

    if all_transactions:
        all_txns_df = pd.DataFrame(all_transactions)
        all_txns_ws = wb.create_sheet(title="All Transactions", index=0)

        # Add title
        all_txns_ws['A1'] = "All Transactions (Consolidated)"
        all_txns_ws['A1'].font = Font(bold=True, size=14)

        # Add summary stats
        all_txns_ws['A3'] = f"Total Transactions: {len(all_transactions):,}"
        all_txns_ws['C3'] = f"Total Money In: £{all_txns_df['Money In'].sum():,.2f}"
        all_txns_ws['E3'] = f"Total Money Out: £{all_txns_df['Money Out'].sum():,.2f}"
        all_txns_ws['A3'].font = Font(bold=True)
        all_txns_ws['C3'].font = Font(bold=True)
        all_txns_ws['E3'].font = Font(bold=True)

        # Add column headers at row 5
        headers = ['Statement', 'Date', 'Description', 'Money In', 'Money Out', 'Balance', 'Confidence %']
        for col_idx, header in enumerate(headers, 1):
            cell = all_txns_ws.cell(row=5, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')

        # Add data starting at row 6
        for row_idx, row_data in enumerate(dataframe_to_rows(all_txns_df, index=False, header=False), 6):
            for col_idx, value in enumerate(row_data, 1):
                all_txns_ws.cell(row=row_idx, column=col_idx, value=value)

        # Set column widths
        all_txns_ws.column_dimensions['A'].width = 35  # Statement name
        all_txns_ws.column_dimensions['B'].width = 12  # Date
        all_txns_ws.column_dimensions['C'].width = 50  # Description
        all_txns_ws.column_dimensions['D'].width = 12  # Money In
        all_txns_ws.column_dimensions['E'].width = 12  # Money Out
        all_txns_ws.column_dimensions['F'].width = 12  # Balance
        all_txns_ws.column_dimensions['G'].width = 12  # Confidence

        # Freeze header rows
        all_txns_ws.freeze_panes = 'A6'

        logger.info(f"  ✓ Consolidated {len(all_transactions):,} transactions from {len(results)} statements")

    # Create summary sheet
    logger.info("Creating summary sheet...")

    summary_ws = wb.create_sheet(title="Summary", index=1)
    summary_ws['A1'] = "Batch Extraction Summary"
    summary_ws['A1'].font = Font(bold=True, size=14)

    summary_ws['A3'] = f"Total Files: {len(pdf_files)}"
    summary_ws['A4'] = f"Successful: {sum(1 for r in results if r['success'])}"
    summary_ws['A5'] = f"Failed: {sum(1 for r in results if not r['success'])}"
    summary_ws['A6'] = f"Total Transactions: {sum(r['transactions'] for r in results)}"

    # Add results table
    summary_ws['A8'] = "File"
    summary_ws['B8'] = "Status"
    summary_ws['C8'] = "Transactions"
    summary_ws['D8'] = "Confidence %"
    summary_ws['E8'] = "Reconciled"
    summary_ws['F8'] = "Error"

    for cell in ['A8', 'B8', 'C8', 'D8', 'E8', 'F8']:
        summary_ws[cell].font = Font(bold=True)
        summary_ws[cell].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    for row_idx, result in enumerate(results, 9):
        summary_ws.cell(row=row_idx, column=1, value=result['file'])
        summary_ws.cell(row=row_idx, column=2, value='✓ Success' if result['success'] else '✗ Failed')
        summary_ws.cell(row=row_idx, column=3, value=result['transactions'] if result['success'] else 0)
        summary_ws.cell(row=row_idx, column=4, value=result.get('confidence', 0))
        summary_ws.cell(row=row_idx, column=5, value='✓' if result.get('reconciled') else '✗')
        summary_ws.cell(row=row_idx, column=6, value=result.get('error', ''))

    summary_ws.column_dimensions['A'].width = 50
    summary_ws.column_dimensions['B'].width = 15
    summary_ws.column_dimensions['C'].width = 15
    summary_ws.column_dimensions['D'].width = 15
    summary_ws.column_dimensions['E'].width = 12
    summary_ws.column_dimensions['F'].width = 50

    # Save workbook
    wb.save(output)
    logger.info(f"\n✓ Batch extraction complete!")
    logger.info(f"  Output: {output}")
    logger.info(f"  Total sheets: {len(wb.sheetnames)}")
    logger.info("=" * 80)

    return True


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python batch_extract.py <folder_path> <output_excel>")
        print("Example: python batch_extract.py statements/Proudfoot proudfoot_all.xlsx")
        sys.exit(1)

    folder_path = sys.argv[1]
    output_path = sys.argv[2]

    success = batch_extract(folder_path, output_path)
    sys.exit(0 if success else 1)
