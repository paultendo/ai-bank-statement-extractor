#!/usr/bin/env python3
"""Quick smoke runner for validated bank PDFs.

Runs the CLI extractor against a curated list of fixtures and checks:
- Exit code / CLI success
- Transaction count > 0
- Balance reconciliation flag
- Optional bank-specific assertions (e.g., pot summaries, txn counts)
"""
from __future__ import annotations

import json
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
STATEMENTS = ROOT / "statements"
OUTPUT = ROOT / "output"


@dataclass
class SmokeCase:
    name: str
    statement: Path
    output: Path
    json_path: Path
    extra_checks: list[Callable[[Path], None]]


def run_cli(case: SmokeCase) -> dict:
    cmd = [
        sys.executable,
        "-m",
        "src.cli",
        "extract",
        str(case.statement),
        "--output",
        str(case.output),
        "--json",
        str(case.json_path),
    ]
    result = subprocess.run(
        cmd,
        cwd=ROOT,
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"CLI failed for {case.name} (exit={result.returncode}):\n{result.stderr or result.stdout}"
        )
    return {
        "stdout": result.stdout,
        "stderr": result.stderr,
    }


def load_result_json(json_path: Path) -> dict:
    if not json_path.exists():
        raise FileNotFoundError(f"Expected result JSON at {json_path}")
    return json.loads(json_path.read_text())


def check_basic_metrics(case: SmokeCase, result_json: dict) -> None:
    if not result_json.get("success"):
        raise AssertionError(f"Smoke failed: {case.name} success flag false")
    if result_json.get("transaction_count", 0) == 0:
        raise AssertionError(f"Smoke failed: {case.name} zero transactions")
    if not result_json.get("balance_reconciled"):
        raise AssertionError(f"Smoke failed: {case.name} balance not reconciled")


def check_tsb_counts(xlsx_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Transactions"]
    rows = max(ws.max_row - 3, 0)
    if rows != 31:
        raise AssertionError(f"TSB expected 31 transactions, found {rows}")
    closing_balance = None
    for row in range(ws.max_row - 1, 1, -1):
        val = ws.cell(row, 5).value
        if isinstance(val, (int, float)):
            closing_balance = float(val)
            break
    if closing_balance is None or abs(closing_balance - 532.24) > 0.01:
        raise AssertionError("TSB closing balance mismatch")


def check_monzo_pots(xlsx_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Pot Summaries" not in wb.sheetnames:
        raise AssertionError("Monzo workbook missing Pot Summaries sheet")
    ws = wb["Pot Summaries"]
    if ws.max_row < 2:
        raise AssertionError("Monzo Pot Summaries sheet is empty")


def check_nationwide_periods(xlsx_path: Path) -> None:
    json_path = xlsx_path.with_suffix(".json")
    result = load_result_json(json_path)
    txns = result.get("transactions") or []

    period_break_indices = [
        idx for idx, txn in enumerate(txns)
        if txn.get("description") == "NATIONWIDE_PERIOD_BREAK"
    ]
    expected_breaks = 13
    if len(period_break_indices) != expected_breaks:
        raise AssertionError(
            f"Nationwide expected {expected_breaks} period breaks, found {len(period_break_indices)}"
        )

    running_balance = None
    periods_checked = 0
    for txn in txns:
        if txn.get("description") == "NATIONWIDE_PERIOD_BREAK":
            running_balance = txn.get("balance")
            periods_checked += 1
            continue

        if running_balance is None:
            raise AssertionError("Nationwide transactions appeared before first period break marker")

        running_balance += txn.get("money_in", 0.0) - txn.get("money_out", 0.0)
        stated_balance = txn.get("balance")
        if stated_balance is None:
            continue
        if abs(running_balance - stated_balance) > 0.02:
            raise AssertionError(
                "Nationwide balance drift detected: "
                f"calc £{running_balance:.2f} vs stated £{stated_balance:.2f} "
                f"on {txn.get('date')} / {txn.get('description')}"
            )

    if periods_checked != expected_breaks:
        raise AssertionError(
            f"Nationwide period markers processed ({periods_checked}) mismatch expected {expected_breaks}"
        )


def make_case(filename: str, extra: list[Callable[[Path], None]] | None = None) -> SmokeCase:
    statement_path = STATEMENTS / filename
    output_path = OUTPUT / (statement_path.stem + "_smoke.xlsx")
    json_path = output_path.with_suffix(".json")
    return SmokeCase(
        name=filename,
        statement=statement_path,
        output=output_path,
        json_path=json_path,
        extra_checks=extra or [],
    )


CASES = [
    make_case("Statements 1.pdf"),
    make_case("Statements 2.pdf"),
    make_case("Statement 3.pdf"),
    make_case("Lillian Gyamfi Halifax Statement Dec 24.pdf"),
    make_case("Lillian Gyamfi Halifax Statement Jan 25.pdf"),
    make_case("HSBC Combined Statements for Myah Wright.pdf"),
    make_case("Lloyds - Deborah Prime.pdf"),
    make_case("Proudfoot/Statement 31-MAY-24 AC 33688186  02065716.pdf"),
    make_case("CurrentAccountStatement_08022024.pdf"),
    make_case(
        "TSB Savings account - Mark Wilcox.pdf",
        extra=[check_tsb_counts],
    ),
    make_case(
        "monzo-bidmead.pdf",
        extra=[check_monzo_pots],
    ),
    make_case(
        "Marsh Bankstatements up to April 2024.pdf",
        extra=[check_nationwide_periods],
    ),
]


def main() -> None:
    failures = []
    for case in CASES:
        try:
            run_cli(case)
            result_json = load_result_json(case.json_path)
            check_basic_metrics(case, result_json)
            for extra_check in case.extra_checks:
                extra_check(case.output)
            print(f"[OK] {case.name}")
        except Exception as exc:  # noqa: BLE001
            failures.append((case.name, exc))
            print(f"[FAIL] {case.name}: {exc}")

    if failures:
        print("\nSmoke failures:")
        for name, exc in failures:
            print(f" - {name}: {exc}")
        sys.exit(1)


if __name__ == "__main__":
    main()
