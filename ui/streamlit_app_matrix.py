"""
Bank Statement Extractor - Streamlit UI

A beautiful, user-friendly interface for extracting bank statement data.
Designed for Fifty Six Law legal team via AILEX.
"""

import streamlit as st
import pandas as pd
import sys
import os
from pathlib import Path
from datetime import datetime
import base64
from io import BytesIO

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.pipeline import ExtractionPipeline
from src.config import settings

# Page configuration
st.set_page_config(
    page_title="Bank Statement Extractor",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for sleek, modern Apple-inspired UI
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Space+Mono:wght@400;600&display=swap');

    /* FORCE DARK BACKGROUND EVERYWHERE */
    .stApp {
        background-color: #000000 !important;
    }
    .main {
        background-color: #000000 !important;
        color: #00ff41 !important;
    }
    [data-testid="stSidebar"] {
        background-color: #000000 !important;
        border-right: 2px solid #00ff41 !important;
    }
    [data-testid="stSidebar"] > div {
        background-color: #000000 !important;
    }

    /* Matrix green text everywhere */
    body, p, div, span, label, .stMarkdown {
        color: #00ff41 !important;
        font-family: 'Space Mono', monospace !important;
    }

    /* Headers with neon glow */
    h1, h2, h3 {
        color: #00ff41 !important;
        font-family: 'Space Mono', monospace !important;
        text-shadow: 0 0 20px rgba(0, 255, 65, 0.8), 0 0 40px rgba(0, 255, 65, 0.5) !important;
        letter-spacing: 2px !important;
        text-transform: uppercase !important;
    }

    /* File uploader with neon border */
    [data-testid="stFileUploader"] {
        background-color: rgba(0, 255, 65, 0.05) !important;
        border: 2px dashed #00ff41 !important;
        border-radius: 10px !important;
        padding: 2rem !important;
        box-shadow: 0 0 30px rgba(0, 255, 65, 0.3) !important;
    }
    [data-testid="stFileUploader"] label {
        color: #00ff41 !important;
    }

    /* Metric cards with glow */
    .metric-card {
        background: rgba(0, 20, 10, 0.8) !important;
        border: 2px solid #00ff41 !important;
        border-radius: 10px !important;
        padding: 1.5rem !important;
        box-shadow: 0 0 40px rgba(0, 255, 65, 0.4), inset 0 0 20px rgba(0, 255, 65, 0.1) !important;
        animation: pulse 2s ease-in-out infinite !important;
    }
    .success-card {
        background: rgba(0, 255, 65, 0.2) !important;
        border: 2px solid #00ff41 !important;
        border-radius: 10px !important;
        padding: 1.5rem !important;
        box-shadow: 0 0 60px rgba(0, 255, 65, 0.6) !important;
    }
    .warning-card {
        background: rgba(255, 0, 0, 0.2) !important;
        border: 2px solid #ff0055 !important;
        border-radius: 10px !important;
        padding: 1.5rem !important;
        box-shadow: 0 0 60px rgba(255, 0, 85, 0.6) !important;
    }

    @keyframes pulse {
        0%, 100% {
            box-shadow: 0 0 30px rgba(0, 255, 65, 0.3), inset 0 0 15px rgba(0, 255, 65, 0.1);
        }
        50% {
            box-shadow: 0 0 60px rgba(0, 255, 65, 0.6), inset 0 0 30px rgba(0, 255, 65, 0.2);
        }
    }

    /* Neon buttons */
    .stButton>button {
        width: 100% !important;
        background: rgba(0, 255, 65, 0.1) !important;
        color: #00ff41 !important;
        border: 3px solid #00ff41 !important;
        border-radius: 0px !important;
        padding: 1rem 2rem !important;
        font-family: 'Space Mono', monospace !important;
        font-weight: 700 !important;
        font-size: 1.1rem !important;
        text-transform: uppercase !important;
        letter-spacing: 3px !important;
        box-shadow: 0 0 40px rgba(0, 255, 65, 0.5), inset 0 0 20px rgba(0, 255, 65, 0.2) !important;
        transition: all 0.3s ease !important;
    }
    .stButton>button:hover {
        background: rgba(0, 255, 65, 0.3) !important;
        box-shadow: 0 0 80px rgba(0, 255, 65, 0.8), inset 0 0 40px rgba(0, 255, 65, 0.4) !important;
        transform: translateY(-2px) !important;
    }

    /* Selectbox with neon style */
    .stSelectbox > div > div {
        background-color: rgba(0, 20, 10, 0.8) !important;
        border: 2px solid #00ff41 !important;
        color: #00ff41 !important;
        border-radius: 0px !important;
        box-shadow: 0 0 20px rgba(0, 255, 65, 0.3) !important;
    }

    /* Metric styling */
    [data-testid="stMetric"] {
        background-color: rgba(0, 20, 10, 0.8) !important;
        border: 1px solid #00ff41 !important;
        border-radius: 5px !important;
        padding: 1rem !important;
        box-shadow: 0 0 20px rgba(0, 255, 65, 0.3) !important;
    }
    [data-testid="stMetricLabel"] {
        color: #00ff41 !important;
    }
    [data-testid="stMetricValue"] {
        color: #00ff41 !important;
        text-shadow: 0 0 10px rgba(0, 255, 65, 0.8) !important;
    }

    /* DataFrame with Matrix styling */
    .dataframe {
        background-color: rgba(0, 20, 10, 0.8) !important;
        border: 2px solid #00ff41 !important;
        font-family: 'Space Mono', monospace !important;
        color: #00ff41 !important;
    }
    .dataframe thead tr th {
        background-color: rgba(0, 255, 65, 0.2) !important;
        color: #00ff41 !important;
        border: 1px solid #00ff41 !important;
        font-weight: 700 !important;
        text-transform: uppercase !important;
    }
    .dataframe tbody tr td {
        background-color: rgba(0, 10, 5, 0.9) !important;
        color: #00ff41 !important;
        border: 1px solid rgba(0, 255, 65, 0.3) !important;
    }

    /* Bank badges */
    .bank-badge {
        background: rgba(0, 255, 65, 0.2) !important;
        border: 2px solid #00ff41 !important;
        color: #00ff41 !important;
        padding: 0.4rem 1rem !important;
        border-radius: 0px !important;
        font-family: 'Space Mono', monospace !important;
        font-weight: 700 !important;
        box-shadow: 0 0 20px rgba(0, 255, 65, 0.4) !important;
        text-transform: uppercase !important;
        letter-spacing: 1px !important;
    }

    /* Success/warning alerts */
    .stAlert {
        background-color: rgba(0, 255, 65, 0.1) !important;
        border: 2px solid #00ff41 !important;
        border-radius: 0px !important;
        color: #00ff41 !important;
    }

    /* Download button */
    .stDownloadButton>button {
        background: rgba(0, 255, 65, 0.1) !important;
        color: #00ff41 !important;
        border: 2px solid #00ff41 !important;
        font-family: 'Space Mono', monospace !important;
        text-transform: uppercase !important;
        box-shadow: 0 0 30px rgba(0, 255, 65, 0.4) !important;
    }

    /* Info boxes */
    .stInfo {
        background: rgba(0, 100, 255, 0.1) !important;
        border: 2px solid #00ccff !important;
        color: #00ccff !important;
    }

    /* Remove default padding */
    .block-container {
        padding-top: 2rem !important;
    }
    </style>
""", unsafe_allow_html=True)

# Initialize session state
if 'extraction_result' not in st.session_state:
    st.session_state.extraction_result = None
if 'uploaded_file_data' not in st.session_state:
    st.session_state.uploaded_file_data = None

def get_bank_color(bank: str) -> str:
    """Return color for bank badge"""
    colors = {
        'barclays': '#00AEEF',
        'hsbc': '#DB0011',
        'lloyds': '#006F4E',
        'natwest': '#5A287B',
        'rbs': '#003A70',
        'santander': '#EC0000',
        'nationwide': '#0057A8',
        'tsb': '#0066B3',
        'monzo': '#14233C'
    }
    return colors.get(bank.lower(), '#667eea')

def create_bank_badge(bank: str) -> str:
    """Create HTML for bank badge"""
    color = get_bank_color(bank)
    return f'<span class="bank-badge" style="background-color: {color}; color: white;">{bank.upper()}</span>'

def format_currency(amount: float) -> str:
    """Format amount as currency"""
    if pd.isna(amount):
        return "-"
    return f"£{amount:,.2f}"

def main():
    # Header
    st.markdown("""
        <div style='text-align: center; padding: 2rem 0;'>
            <h1 style='color: #2c3e50; font-size: 3rem; margin-bottom: 0;'>
                🏦 Bank Statement Extractor
            </h1>
            <p style='color: #7f8c8d; font-size: 1.2rem; margin-top: 0.5rem;'>
                Automated data extraction for legal evidence
            </p>
        </div>
    """, unsafe_allow_html=True)

    st.markdown("---")

    # Sidebar
    with st.sidebar:
        st.markdown("### ⚙️ Settings")

        # Bank selection
        supported_banks = [
            'Auto-detect',
            'Barclays',
            'HSBC',
            'Lloyds',
            'NatWest',
            'RBS',
            'Santander',
            'Nationwide',
            'TSB',
            'Monzo'
        ]

        selected_bank = st.selectbox(
            "Select Bank (or auto-detect)",
            supported_banks,
            help="Choose the bank or let the system auto-detect"
        )

        # Output format
        output_format = st.selectbox(
            "Output Format",
            ['Excel (.xlsx)', 'CSV (.csv)'],
            help="Choose export format"
        )

        st.markdown("---")

        # Supported banks display
        st.markdown("### 🏦 Supported Banks")
        for bank in supported_banks[1:]:  # Skip auto-detect
            st.markdown(create_bank_badge(bank), unsafe_allow_html=True)

        st.markdown("---")

        # Info
        st.markdown("""
            ### 📋 Features
            - ✅ PDF & Image support
            - ✅ Auto bank detection
            - ✅ Balance validation
            - ✅ High accuracy OCR
            - ✅ Excel/CSV export
            - ✅ Audit logging

            ### 🔒 Security
            All processing is done locally.
            No data is sent to external servers.
        """)

    # Main content area
    col1, col2 = st.columns([1, 1])

    with col1:
        st.markdown("### 📤 Upload Statement")

        # File uploader
        uploaded_file = st.file_uploader(
            "Choose a bank statement file",
            type=['pdf', 'png', 'jpg', 'jpeg'],
            help="Upload a PDF or image of a bank statement",
            label_visibility="collapsed"
        )

        if uploaded_file:
            # Save uploaded file data
            st.session_state.uploaded_file_data = uploaded_file.read()
            uploaded_file.seek(0)  # Reset file pointer

            # Display file info
            st.success(f"✅ Uploaded: **{uploaded_file.name}** ({len(st.session_state.uploaded_file_data) / 1024:.1f} KB)")

            # Display file preview if PDF
            if uploaded_file.name.lower().endswith('.pdf'):
                st.info("📄 PDF preview available after processing")
            else:
                # Show image preview
                st.image(uploaded_file, caption=uploaded_file.name, use_column_width=True)

    with col2:
        st.markdown("### 🚀 Process Statement")

        if uploaded_file:
            if st.button("🔍 Extract Data", type="primary"):
                # Save file temporarily
                temp_path = Path(f"/tmp/{uploaded_file.name}")
                with open(temp_path, "wb") as f:
                    f.write(st.session_state.uploaded_file_data)

                # Process with progress indicator
                with st.spinner("🔄 Processing statement..."):
                    try:
                        # Initialize pipeline
                        pipeline = ExtractionPipeline()

                        # Determine bank parameter
                        bank_param = None if selected_bank == 'Auto-detect' else selected_bank.lower()

                        # Process statement
                        result = pipeline.process(temp_path, bank_name=bank_param)

                        # Store result in session state
                        st.session_state.extraction_result = result

                        st.success("✅ Processing complete!")
                        st.rerun()

                    except Exception as e:
                        st.error(f"❌ Extraction failed: {str(e)}")
                        st.exception(e)
                    finally:
                        # Cleanup temp file
                        if os.path.exists(temp_path):
                            os.remove(temp_path)
        else:
            st.info("👆 Upload a statement file to begin")

    # Results section
    if st.session_state.extraction_result:
        result = st.session_state.extraction_result

        st.markdown("---")
        st.markdown("## 📊 Extraction Results")

        # Metrics row
        metric_cols = st.columns(4)

        with metric_cols[0]:
            st.markdown(f"""
                <div class="metric-card">
                    <h3 style="margin: 0; font-size: 1.1rem;">Bank Detected</h3>
                    <p style="margin: 0.5rem 0 0 0; font-size: 1.8rem; font-weight: bold;">
                        {result.statement.bank_name.upper() if result.statement.bank_name else 'Unknown'}
                    </p>
                </div>
            """, unsafe_allow_html=True)

        with metric_cols[1]:
            st.markdown(f"""
                <div class="metric-card">
                    <h3 style="margin: 0; font-size: 1.1rem;">Transactions</h3>
                    <p style="margin: 0.5rem 0 0 0; font-size: 1.8rem; font-weight: bold;">
                        {len(result.transactions)}
                    </p>
                </div>
            """, unsafe_allow_html=True)

        with metric_cols[2]:
            confidence = result.confidence_score
            confidence_class = "success-card" if confidence > 90 else "warning-card" if confidence > 70 else "metric-card"
            st.markdown(f"""
                <div class="{confidence_class}">
                    <h3 style="margin: 0; font-size: 1.1rem;">Confidence</h3>
                    <p style="margin: 0.5rem 0 0 0; font-size: 1.8rem; font-weight: bold;">
                        {confidence:.1f}%
                    </p>
                </div>
            """, unsafe_allow_html=True)

        with metric_cols[3]:
            reconciled = result.balance_reconciled
            reconcile_class = "success-card" if reconciled else "warning-card"
            reconcile_icon = "✅" if reconciled else "⚠️"
            st.markdown(f"""
                <div class="{reconcile_class}">
                    <h3 style="margin: 0; font-size: 1.1rem;">Balance Check</h3>
                    <p style="margin: 0.5rem 0 0 0; font-size: 1.8rem; font-weight: bold;">
                        {reconcile_icon} {'Pass' if reconciled else 'Review'}
                    </p>
                </div>
            """, unsafe_allow_html=True)

        # Statement details
        st.markdown("### 📋 Statement Details")
        detail_cols = st.columns(3)

        with detail_cols[0]:
            st.metric("Account Number", result.statement.account_number or "N/A")

        with detail_cols[1]:
            if result.statement.statement_start_date and result.statement.statement_end_date:
                period = f"{result.statement.statement_start_date.strftime('%d/%m/%Y')} - {result.statement.statement_end_date.strftime('%d/%m/%Y')}"
            else:
                period = "N/A"
            st.metric("Statement Period", period)

        with detail_cols[2]:
            opening = format_currency(result.statement.opening_balance)
            closing = format_currency(result.statement.closing_balance)
            st.metric("Balance", f"{opening} → {closing}")

        # Transactions table
        st.markdown("### 💰 Transactions")

        # Convert to DataFrame
        df_data = []
        for txn in result.transactions:
            df_data.append({
                'Date': txn.date.strftime('%d/%m/%Y') if txn.date else '',
                'Description': txn.description,
                'Money In': txn.money_in if txn.money_in > 0 else None,
                'Money Out': txn.money_out if txn.money_out > 0 else None,
                'Balance': txn.balance,
                'Type': txn.transaction_type.value if txn.transaction_type else '',
                'Confidence': txn.confidence if txn.confidence else 0.0
            })

        df = pd.DataFrame(df_data)

        # Display with formatting
        st.dataframe(
            df.style.format({
                'Money In': lambda x: format_currency(x) if pd.notna(x) else '-',
                'Money Out': lambda x: format_currency(x) if pd.notna(x) else '-',
                'Balance': lambda x: format_currency(x) if pd.notna(x) else '-',
                'Confidence': lambda x: f"{x:.1f}%" if pd.notna(x) and x > 0 else '-'
            }).background_gradient(
                subset=['Confidence'],
                cmap='RdYlGn',
                vmin=70,
                vmax=100
            ),
            use_container_width=True,
            height=400
        )

        # Summary statistics
        st.markdown("### 📈 Summary")
        summary_cols = st.columns(3)

        total_in = sum(txn.money_in for txn in result.transactions)
        total_out = sum(txn.money_out for txn in result.transactions)
        net_change = total_in - total_out

        with summary_cols[0]:
            st.metric("Total Money In", format_currency(total_in), delta=None)

        with summary_cols[1]:
            st.metric("Total Money Out", format_currency(total_out), delta=None)

        with summary_cols[2]:
            delta_color = "normal" if net_change >= 0 else "inverse"
            st.metric("Net Change", format_currency(net_change), delta=format_currency(net_change), delta_color=delta_color)

        # Validation messages
        if result.warnings:
            st.markdown("### ⚠️ Validation Warnings")
            for msg in result.warnings:
                st.warning(msg)

        # Export section
        st.markdown("### 💾 Export Data")
        export_cols = st.columns(2)

        with export_cols[0]:
            # Generate Excel file
            output = BytesIO()

            # Create Excel file with formatting
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Transactions"

            # Headers
            headers = ['Date', 'Description', 'Money In', 'Money Out', 'Balance', 'Type']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Data
            for row, txn in enumerate(result.transactions, 2):
                ws.cell(row=row, column=1, value=txn.date.strftime('%d/%m/%Y') if txn.date else '')
                ws.cell(row=row, column=2, value=txn.description)
                ws.cell(row=row, column=3, value=txn.money_in if txn.money_in > 0 else None)
                ws.cell(row=row, column=4, value=txn.money_out if txn.money_out > 0 else None)
                ws.cell(row=row, column=5, value=txn.balance)
                ws.cell(row=row, column=6, value=txn.transaction_type.value if txn.transaction_type else '')

            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 20

            wb.save(output)
            excel_data = output.getvalue()

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            bank_name = result.statement.bank_name or 'statement'
            filename = f"{bank_name}_extracted_{timestamp}.xlsx"

            st.download_button(
                label="📥 Download Excel",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with export_cols[1]:
            # Generate CSV
            csv_data = df.to_csv(index=False)
            csv_filename = f"{bank_name}_extracted_{timestamp}.csv"

            st.download_button(
                label="📥 Download CSV",
                data=csv_data,
                file_name=csv_filename,
                mime="text/csv",
                use_container_width=True
            )

    # Footer
    st.markdown("---")
    st.markdown("""
        <div style='text-align: center; color: #7f8c8d; padding: 2rem 0;'>
            <p>Built for <strong>instant bank statement extraction</strong></p>
            <p style='font-size: 0.9rem;'>🔒 All processing is local. Your data never leaves this system.</p>
        </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
