"""
Bank Statement Extractor - Streamlit UI
Apple-inspired liquid glass design
"""

import streamlit as st
import pandas as pd
import sys
import os
from pathlib import Path
from datetime import datetime
from io import BytesIO

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.pipeline import ExtractionPipeline

# Page config
st.set_page_config(
    page_title="Bank Statement Extractor",
    page_icon="💎",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Apple-inspired liquid glass CSS
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Space+Mono:wght@500;700&display=swap');

    /* Global dark theme */
    .stApp {
        background: linear-gradient(180deg, #0a0a0a 0%, #1a1a1a 100%);
    }
    .main {
        background: transparent;
        padding: 2rem 3rem;
    }

    /* Typography */
    * {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
    }

    .monospace {
        font-family: 'Space Mono', monospace !important;
    }

    h1, h2, h3 {
        color: #ffffff !important;
        font-weight: 700 !important;
        letter-spacing: -0.02em !important;
    }

    p, div, span, label {
        color: rgba(255, 255, 255, 0.85) !important;
    }

    /* Liquid glass cards */
    .glass-card {
        background: rgba(255, 255, 255, 0.05);
        border: 1px solid rgba(255, 255, 255, 0.1);
        border-radius: 16px;
        padding: 1.5rem;
        backdrop-filter: blur(20px) saturate(180%);
        -webkit-backdrop-filter: blur(20px) saturate(180%);
        box-shadow:
            0 8px 32px 0 rgba(0, 0, 0, 0.37),
            inset 0 1px 0 0 rgba(255, 255, 255, 0.1);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
    }

    .glass-card:hover {
        background: rgba(255, 255, 255, 0.08);
        border-color: rgba(255, 255, 255, 0.2);
        transform: translateY(-2px);
        box-shadow:
            0 12px 48px 0 rgba(0, 0, 0, 0.5),
            inset 0 1px 0 0 rgba(255, 255, 255, 0.15);
    }

    /* Sidebar glass effect */
    [data-testid="stSidebar"] {
        background: rgba(18, 18, 18, 0.8) !important;
        border-right: 1px solid rgba(255, 255, 255, 0.1) !important;
        backdrop-filter: blur(40px) saturate(180%);
    }

    /* Buttons with glass morphism */
    .stButton>button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
        border: none !important;
        border-radius: 12px !important;
        color: white !important;
        font-weight: 600 !important;
        padding: 0.75rem 2rem !important;
        box-shadow:
            0 4px 24px rgba(102, 126, 234, 0.4),
            inset 0 1px 0 rgba(255, 255, 255, 0.2) !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
    }

    .stButton>button:hover {
        transform: translateY(-2px) scale(1.02) !important;
        box-shadow:
            0 8px 32px rgba(102, 126, 234, 0.6),
            inset 0 1px 0 rgba(255, 255, 255, 0.3) !important;
    }

    /* File uploader glass */
    [data-testid="stFileUploader"] {
        background: rgba(255, 255, 255, 0.03);
        border: 2px dashed rgba(255, 255, 255, 0.2);
        border-radius: 16px;
        padding: 2rem;
        backdrop-filter: blur(10px);
    }

    /* Metrics with glass effect */
    [data-testid="stMetric"] {
        background: rgba(255, 255, 255, 0.05);
        border: 1px solid rgba(255, 255, 255, 0.1);
        border-radius: 12px;
        padding: 1.25rem;
        backdrop-filter: blur(20px);
        box-shadow: 0 4px 16px rgba(0, 0, 0, 0.3);
    }

    [data-testid="stMetricLabel"] {
        color: rgba(255, 255, 255, 0.6) !important;
        font-size: 0.875rem !important;
        font-weight: 500 !important;
        text-transform: uppercase !important;
        letter-spacing: 0.05em !important;
    }

    [data-testid="stMetricValue"] {
        color: #ffffff !important;
        font-size: 2rem !important;
        font-weight: 700 !important;
    }

    /* Success metric - green glass */
    .metric-success {
        background: linear-gradient(135deg, rgba(16, 185, 129, 0.1) 0%, rgba(5, 150, 105, 0.1) 100%);
        border-color: rgba(16, 185, 129, 0.3);
        box-shadow: 0 4px 24px rgba(16, 185, 129, 0.2);
    }

    /* Warning metric - amber glass */
    .metric-warning {
        background: linear-gradient(135deg, rgba(245, 158, 11, 0.1) 0%, rgba(217, 119, 6, 0.1) 100%);
        border-color: rgba(245, 158, 11, 0.3);
        box-shadow: 0 4px 24px rgba(245, 158, 11, 0.2);
    }

    /* Select box glass */
    .stSelectbox > div > div {
        background: rgba(255, 255, 255, 0.05) !important;
        border: 1px solid rgba(255, 255, 255, 0.1) !important;
        border-radius: 10px !important;
        backdrop-filter: blur(10px) !important;
        color: white !important;
    }

    /* DataFrame with premium styling */
    .dataframe {
        background: rgba(255, 255, 255, 0.03) !important;
        border: 1px solid rgba(255, 255, 255, 0.1) !important;
        border-radius: 12px !important;
        font-family: 'Space Mono', monospace !important;
        font-size: 0.9rem !important;
    }

    .dataframe thead tr th {
        background: rgba(255, 255, 255, 0.08) !important;
        color: rgba(255, 255, 255, 0.9) !important;
        font-weight: 600 !important;
        text-transform: uppercase !important;
        font-size: 0.75rem !important;
        letter-spacing: 0.05em !important;
        padding: 1rem !important;
        border-bottom: 1px solid rgba(255, 255, 255, 0.1) !important;
    }

    .dataframe tbody tr td {
        color: rgba(255, 255, 255, 0.85) !important;
        padding: 0.875rem 1rem !important;
        border-bottom: 1px solid rgba(255, 255, 255, 0.05) !important;
    }

    .dataframe tbody tr:hover {
        background: rgba(255, 255, 255, 0.05) !important;
    }

    /* Download button */
    .stDownloadButton>button {
        background: rgba(255, 255, 255, 0.1) !important;
        border: 1px solid rgba(255, 255, 255, 0.2) !important;
        border-radius: 10px !important;
        color: white !important;
        font-weight: 600 !important;
        padding: 0.75rem 1.5rem !important;
        backdrop-filter: blur(10px) !important;
        transition: all 0.3s ease !important;
    }

    .stDownloadButton>button:hover {
        background: rgba(255, 255, 255, 0.15) !important;
        border-color: rgba(255, 255, 255, 0.3) !important;
        transform: translateY(-2px) !important;
    }

    /* Alerts with glass effect */
    .stAlert {
        background: rgba(59, 130, 246, 0.1) !important;
        border: 1px solid rgba(59, 130, 246, 0.3) !important;
        border-radius: 12px !important;
        color: rgba(147, 197, 253, 1) !important;
        backdrop-filter: blur(10px) !important;
    }

    /* Success alert */
    .stSuccess {
        background: rgba(16, 185, 129, 0.1) !important;
        border-color: rgba(16, 185, 129, 0.3) !important;
        color: rgba(110, 231, 183, 1) !important;
    }

    /* Warning alert */
    .stWarning {
        background: rgba(245, 158, 11, 0.1) !important;
        border-color: rgba(245, 158, 11, 0.3) !important;
        color: rgba(252, 211, 77, 1) !important;
    }

    /* Error alert */
    .stError {
        background: rgba(239, 68, 68, 0.1) !important;
        border-color: rgba(239, 68, 68, 0.3) !important;
        color: rgba(252, 165, 165, 1) !important;
    }

    /* Info boxes */
    .stInfo {
        background: rgba(139, 92, 246, 0.1) !important;
        border: 1px solid rgba(139, 92, 246, 0.3) !important;
        border-radius: 12px !important;
        color: rgba(196, 181, 253, 1) !important;
    }

    /* Remove default margins */
    .block-container {
        padding-top: 2rem !important;
        max-width: 1400px !important;
    }

    /* Smooth scrollbar */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }

    ::-webkit-scrollbar-track {
        background: rgba(255, 255, 255, 0.05);
    }

    ::-webkit-scrollbar-thumb {
        background: rgba(255, 255, 255, 0.2);
        border-radius: 4px;
    }

    ::-webkit-scrollbar-thumb:hover {
        background: rgba(255, 255, 255, 0.3);
    }
    </style>
""", unsafe_allow_html=True)

# Session state
if 'extraction_result' not in st.session_state:
    st.session_state.extraction_result = None
if 'uploaded_file_data' not in st.session_state:
    st.session_state.uploaded_file_data = None

def format_currency(amount: float) -> str:
    if pd.isna(amount):
        return "-"
    return f"£{amount:,.2f}"

def main():
    # Header
    st.markdown("""
        <div style='text-align: center; padding: 3rem 0 2rem 0;'>
            <h1 style='font-size: 3.5rem; margin-bottom: 0.5rem; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); -webkit-background-clip: text; -webkit-text-fill-color: transparent;'>
                💎 Bank Statement Extractor
            </h1>
            <p style='color: rgba(255, 255, 255, 0.6); font-size: 1.1rem; font-weight: 500;'>
                Automated data extraction for legal evidence
            </p>
        </div>
    """, unsafe_allow_html=True)

    # Sidebar
    with st.sidebar:
        st.markdown("### ⚙️ Configuration")

        supported_banks = [
            'Auto-detect', 'Barclays', 'HSBC', 'Lloyds', 'NatWest',
            'RBS', 'Santander', 'Nationwide', 'TSB', 'Monzo'
        ]

        selected_bank = st.selectbox("Bank", supported_banks)
        output_format = st.selectbox("Format", ['Excel (.xlsx)', 'CSV (.csv)'])

        st.markdown("---")
        st.markdown("### 🏦 Supported Banks")
        for bank in supported_banks[1:]:
            st.markdown(f"`{bank}`")

        st.markdown("---")
        st.markdown("""
            ### 🔒 Privacy
            All processing is **100% local**.
            Your data never leaves this system.
        """)

    # Main content
    col1, col2 = st.columns([1, 1])

    with col1:
        st.markdown("### 📤 Upload Statement")
        uploaded_file = st.file_uploader(
            "Drag and drop or browse",
            type=['pdf', 'png', 'jpg', 'jpeg'],
            label_visibility="collapsed"
        )

        if uploaded_file:
            st.session_state.uploaded_file_data = uploaded_file.read()
            uploaded_file.seek(0)
            st.success(f"✅ {uploaded_file.name} ({len(st.session_state.uploaded_file_data) / 1024:.1f} KB)")

    with col2:
        st.markdown("### 🚀 Process")
        if uploaded_file:
            if st.button("Extract Data", type="primary"):
                temp_path = Path(f"/tmp/{uploaded_file.name}")
                with open(temp_path, "wb") as f:
                    f.write(st.session_state.uploaded_file_data)

                with st.spinner("Processing..."):
                    try:
                        pipeline = ExtractionPipeline()
                        bank_param = None if selected_bank == 'Auto-detect' else selected_bank.lower()
                        result = pipeline.process(temp_path, bank_name=bank_param)
                        st.session_state.extraction_result = result
                        st.success("✅ Complete!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
                    finally:
                        if os.path.exists(temp_path):
                            os.remove(temp_path)
        else:
            st.info("👆 Upload a statement to begin")

    # Results
    if st.session_state.extraction_result:
        result = st.session_state.extraction_result

        st.markdown("---")
        st.markdown("## 📊 Results")

        # Metrics
        cols = st.columns(4)
        with cols[0]:
            st.metric("Bank", result.statement.bank_name.upper() if result.statement.bank_name else 'Unknown')
        with cols[1]:
            st.metric("Transactions", len(result.transactions))
        with cols[2]:
            confidence = result.confidence_score
            st.metric("Confidence", f"{confidence:.1f}%")
        with cols[3]:
            st.metric("Status", "✅ Reconciled" if result.balance_reconciled else "⚠️ Review")

        # Transactions table
        st.markdown("### 💰 Transactions")
        df_data = []
        for txn in result.transactions:
            df_data.append({
                'Date': txn.date.strftime('%d/%m/%Y') if txn.date else '',
                'Description': txn.description,
                'In': txn.money_in if txn.money_in > 0 else None,
                'Out': txn.money_out if txn.money_out > 0 else None,
                'Balance': txn.balance,
                'Confidence': txn.confidence if txn.confidence else 0.0
            })

        df = pd.DataFrame(df_data)
        st.dataframe(
            df.style.format({
                'In': lambda x: format_currency(x) if pd.notna(x) else '-',
                'Out': lambda x: format_currency(x) if pd.notna(x) else '-',
                'Balance': lambda x: format_currency(x) if pd.notna(x) else '-',
                'Confidence': lambda x: f"{x:.1f}%" if pd.notna(x) and x > 0 else '-'
            }).background_gradient(
                subset=['Confidence'],
                cmap='RdYlGn',
                vmin=70,
                vmax=100
            ),
            use_container_width=True,
            height=400
        )

        # Summary
        st.markdown("### 📈 Summary")
        sum_cols = st.columns(3)
        total_in = sum(txn.money_in for txn in result.transactions)
        total_out = sum(txn.money_out for txn in result.transactions)

        with sum_cols[0]:
            st.metric("Total In", format_currency(total_in))
        with sum_cols[1]:
            st.metric("Total Out", format_currency(total_out))
        with sum_cols[2]:
            st.metric("Net", format_currency(total_in - total_out))

        # Export
        st.markdown("### 💾 Export")
        exp_cols = st.columns(2)

        with exp_cols[0]:
            # Generate Excel
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Transactions"

            headers = ['Date', 'Description', 'Money In', 'Money Out', 'Balance']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)

            for row, txn in enumerate(result.transactions, 2):
                ws.cell(row=row, column=1, value=txn.date.strftime('%d/%m/%Y') if txn.date else '')
                ws.cell(row=row, column=2, value=txn.description)
                ws.cell(row=row, column=3, value=txn.money_in if txn.money_in > 0 else None)
                ws.cell(row=row, column=4, value=txn.money_out if txn.money_out > 0 else None)
                ws.cell(row=row, column=5, value=txn.balance)

            output = BytesIO()
            wb.save(output)

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            bank_name = result.statement.bank_name or 'statement'

            st.download_button(
                "📥 Download Excel",
                data=output.getvalue(),
                file_name=f"{bank_name}_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        with exp_cols[1]:
            csv_data = df.to_csv(index=False)
            st.download_button(
                "📥 Download CSV",
                data=csv_data,
                file_name=f"{bank_name}_{timestamp}.csv",
                mime="text/csv"
            )

if __name__ == "__main__":
    main()
