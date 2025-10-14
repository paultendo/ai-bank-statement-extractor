"""
Bank Statement Extractor - Streamlit UI
Clean, minimal, professional design
"""

import streamlit as st
import pandas as pd
import sys
import os
from pathlib import Path
from datetime import datetime
from io import BytesIO

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.pipeline import ExtractionPipeline
from src.analytics import TransactionAnalyzer

# Page config
st.set_page_config(
    page_title="Bank Statement Extractor",
    page_icon="💎",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom styling - work WITH Streamlit's theme system
st.markdown("""
    <style>
    /* Load custom fonts */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Space+Mono:wght@500;600&display=swap');
    @import url('https://fonts.googleapis.com/css2?family=Material+Symbols+Rounded:opsz,wght,FILL,GRAD@20..48,100..700,0..1,-50..200');

    /* Custom fonts - work with Streamlit theme */
    html, body, [class*="st-"] {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    }

    h1, h2, h3, h4, h5, h6 {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
        font-weight: 700;
        letter-spacing: -0.02em;
    }

    /* Monospace for data */
    code, pre, .monospace {
        font-family: 'Space Mono', 'Courier New', monospace !important;
    }

    /* FIX: Material Symbols for icons ONLY - target specific icon classes */
    .st-emotion-cache-zkd0x0,
    .st-emotion-cache-pd6qx2,
    span[style*="Material Symbols"] {
        font-family: 'Material Symbols Rounded' !important;
        font-feature-settings: 'liga' !important;
        -webkit-font-feature-settings: 'liga' !important;
    }

    /* Buttons */
    .stButton>button {
        border-radius: 8px !important;
        font-weight: 600 !important;
        padding: 0.6rem 1.5rem !important;
        transition: all 0.2s ease !important;
    }

    /* DataFrames */
    .dataframe {
        font-family: 'Space Mono', monospace !important;
        font-size: 0.9rem !important;
    }

    .dataframe thead tr th {
        font-weight: 600 !important;
        text-transform: uppercase !important;
        font-size: 0.75rem !important;
        letter-spacing: 0.05em !important;
        padding: 1rem !important;
        position: relative !important;
        z-index: 1 !important;
    }

    .dataframe tbody tr td {
        padding: 0.875rem 1rem !important;
    }

    /* Fix column sort dropdown visibility */
    [data-testid="stDataFrameResizable"] button {
        z-index: 100 !important;
        position: relative !important;
    }

    /* Fix dropdown menu positioning */
    .stDataFrame [role="menu"] {
        z-index: 1000 !important;
        background: white !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.15) !important;
    }

    /* Metrics */
    [data-testid="stMetricLabel"] {
        font-size: 0.875rem !important;
        font-weight: 500 !important;
        text-transform: uppercase !important;
        letter-spacing: 0.05em !important;
    }

    [data-testid="stMetricValue"] {
        font-size: 1.75rem !important;
        font-weight: 700 !important;
    }

    /* File uploader */
    [data-testid="stFileUploader"] {
        border-radius: 12px !important;
        padding: 1.5rem !important;
    }


    /* Selectbox */
    .stSelectbox > div > div {
        border-radius: 8px !important;
    }

    /* Download buttons */
    .stDownloadButton>button {
        border-radius: 8px !important;
        font-weight: 600 !important;
        padding: 0.6rem 1.5rem !important;
    }

    /* Container spacing */
    .block-container {
        padding-top: 2rem !important;
        max-width: 1400px !important;
    }
    </style>
""", unsafe_allow_html=True)

# Session state
if 'extraction_result' not in st.session_state:
    st.session_state.extraction_result = None
if 'uploaded_file_data' not in st.session_state:
    st.session_state.uploaded_file_data = None

def format_currency(amount: float) -> str:
    if pd.isna(amount):
        return "-"
    return f"£{amount:,.2f}"

def main():
    # Header
    st.markdown("""
        <div style='text-align: center; padding: 2rem 0 1.5rem 0;'>
            <h1 style='font-size: 3rem; margin-bottom: 0.5rem;'>
                💎 Bank Statement Extractor
            </h1>
            <p style='font-size: 1.1rem; opacity: 0.7;'>
                Automated data extraction for legal evidence
            </p>
        </div>
    """, unsafe_allow_html=True)

    # Sidebar
    with st.sidebar:
        st.markdown("### ⚙️ Configuration")

        supported_banks = [
            'Auto-detect', 'Barclays', 'HSBC', 'Lloyds', 'NatWest',
            'RBS', 'Santander', 'Nationwide', 'TSB', 'Monzo'
        ]

        selected_bank = st.selectbox("Bank", supported_banks)
        output_format = st.selectbox("Format", ['Excel (.xlsx)', 'CSV (.csv)'])

        st.markdown("---")
        st.markdown("### 🏦 Supported Banks")
        st.markdown(" · ".join(supported_banks[1:]))

        st.markdown("---")
        st.markdown("""
            ### 🔒 Privacy
            All processing is **100% local**.
            Your data never leaves this system.
        """)

    # Main content
    col1, col2 = st.columns([1, 1])

    with col1:
        st.markdown("### 📤 Upload Statement")
        uploaded_file = st.file_uploader(
            "Drag and drop or browse",
            type=['pdf', 'png', 'jpg', 'jpeg'],
            label_visibility="collapsed"
        )

        if uploaded_file:
            st.session_state.uploaded_file_data = uploaded_file.read()
            uploaded_file.seek(0)
            st.success(f"✅ **{uploaded_file.name}** ({len(st.session_state.uploaded_file_data) / 1024:.1f} KB)")

    with col2:
        st.markdown("### 🚀 Process")
        if uploaded_file:
            if st.button("🔍 Extract Data", use_container_width=True):
                temp_path = Path(f"/tmp/{uploaded_file.name}")
                with open(temp_path, "wb") as f:
                    f.write(st.session_state.uploaded_file_data)

                with st.spinner("Processing statement..."):
                    try:
                        pipeline = ExtractionPipeline()
                        bank_param = None if selected_bank == 'Auto-detect' else selected_bank.lower()
                        result = pipeline.process(temp_path, bank_name=bank_param)
                        st.session_state.extraction_result = result
                        st.success("✅ Extraction complete!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
                    finally:
                        if os.path.exists(temp_path):
                            os.remove(temp_path)
        else:
            st.info("👆 Upload a statement file to begin")

    # Results
    if st.session_state.extraction_result:
        result = st.session_state.extraction_result

        # Check if extraction succeeded
        if not result.success or result.statement is None:
            st.error("❌ Extraction failed")
            if result.error_message:
                st.error(f"Error: {result.error_message}")
            if result.warnings:
                for warning in result.warnings:
                    st.warning(warning)
            return

        st.markdown("---")
        st.markdown("## 📊 Extraction Results")

        # Metrics row
        cols = st.columns(4)

        with cols[0]:
            st.metric(
                "Bank Detected",
                result.statement.bank_name.upper() if result.statement and result.statement.bank_name else 'Unknown'
            )

        with cols[1]:
            st.metric("Transactions", f"{len(result.transactions):,}")

        with cols[2]:
            confidence = result.confidence_score
            st.metric("Confidence", f"{confidence:.1f}%")

        with cols[3]:
            status = "✅ Reconciled" if result.balance_reconciled else "⚠️ Review"
            st.metric("Balance Check", status)

        # Statement details
        st.markdown("### 📋 Statement Details")
        detail_cols = st.columns(3)

        with detail_cols[0]:
            account = result.statement.account_number if result.statement else "N/A"
            st.metric("Account", account or "N/A")

        with detail_cols[1]:
            if result.statement and result.statement.statement_start_date and result.statement.statement_end_date:
                period = f"{result.statement.statement_start_date.strftime('%d/%m/%Y')} - {result.statement.statement_end_date.strftime('%d/%m/%Y')}"
            else:
                period = "N/A"
            st.metric("Period", period)

        with detail_cols[2]:
            if result.statement:
                opening = format_currency(result.statement.opening_balance)
                closing = format_currency(result.statement.closing_balance)
                st.metric("Balance", f"{opening} → {closing}")
            else:
                st.metric("Balance", "N/A")

        # Transactions table
        st.markdown("### 💰 Transactions")

        with st.spinner("Loading transaction table..."):
            df_data = []
            for txn in result.transactions:
                df_data.append({
                    'Date': txn.date.strftime('%d/%m/%Y') if txn.date else '',
                    'Description': txn.description,
                    'Money In': txn.money_in if txn.money_in > 0 else None,
                    'Money Out': txn.money_out if txn.money_out > 0 else None,
                    'Balance': txn.balance,
                    'Type': txn.transaction_type.value if txn.transaction_type else '',
                    'Confidence': txn.confidence if txn.confidence else 0.0
                })

            df = pd.DataFrame(df_data)

            st.dataframe(
                df.style.format({
                    'Money In': lambda x: format_currency(x) if pd.notna(x) else '-',
                    'Money Out': lambda x: format_currency(x) if pd.notna(x) else '-',
                    'Balance': lambda x: format_currency(x) if pd.notna(x) else '-',
                    'Confidence': lambda x: f"{x:.1f}%" if pd.notna(x) and x > 0 else '-'
                }).background_gradient(
                    subset=['Confidence'],
                    cmap='RdYlGn',
                    vmin=70,
                    vmax=100
                ),
                use_container_width=True,
                height=400
            )

        # Summary statistics
        st.markdown("### 📈 Summary")
        sum_cols = st.columns(3)

        total_in = sum(txn.money_in for txn in result.transactions)
        total_out = sum(txn.money_out for txn in result.transactions)
        net = total_in - total_out

        with sum_cols[0]:
            st.metric("Total Money In", format_currency(total_in))

        with sum_cols[1]:
            st.metric("Total Money Out", format_currency(total_out))

        with sum_cols[2]:
            st.metric("Net Change", format_currency(net))

        # Warnings
        if result.warnings:
            st.markdown("### ⚠️ Validation Warnings")
            for msg in result.warnings:
                st.warning(msg)

        # Analytics section
        st.markdown("---")
        st.markdown("## 🔍 Advanced Analytics")
        st.markdown("*Insights for legal case analysis*")

        try:
            analyzer = TransactionAnalyzer(result.transactions)

            # Create tabs for different analyses
            tab1, tab2, tab3, tab4, tab5 = st.tabs([
                "📊 Overview",
                "🚨 Unusual Spending",
                "🎰 Gambling",
                "🔍 Fraud Indicators",
                "🛍️ Lifestyle"
            ])

            with tab1:
                st.markdown("### Financial Overview")

                summary = analyzer.get_summary()

                overview_cols = st.columns(3)
                with overview_cols[0]:
                    st.metric("Transaction Count", f"{summary['total_transactions']:,}")
                    st.metric("Period (days)", f"{summary['date_range']['days']}")

                with overview_cols[1]:
                    st.metric("Average Daily Balance", format_currency(summary['avg_daily_balance']))
                    st.metric("Opening Balance", format_currency(summary['opening_balance']))

                with overview_cols[2]:
                    st.metric("Closing Balance", format_currency(summary['closing_balance']))
                    net_pct = ((summary['net_change'] / summary['opening_balance']) * 100) if summary['opening_balance'] != 0 else 0
                    st.metric("Net Change %", f"{net_pct:+.1f}%")

                # Monthly breakdown
                st.markdown("#### Monthly Breakdown")
                monthly = analyzer.get_monthly_summary()
                if monthly:
                    monthly_df = pd.DataFrame(monthly)
                    monthly_df['Month'] = monthly_df['month']
                    monthly_df['Transactions'] = monthly_df['transaction_count']
                    monthly_df['Money In'] = monthly_df['total_in'].apply(lambda x: format_currency(x))
                    monthly_df['Money Out'] = monthly_df['total_out'].apply(lambda x: format_currency(x))
                    monthly_df['Net'] = monthly_df['net'].apply(lambda x: format_currency(x))
                    monthly_df['Avg Balance'] = monthly_df['avg_balance'].apply(lambda x: format_currency(x))

                    st.dataframe(
                        monthly_df[['Month', 'Transactions', 'Money In', 'Money Out', 'Net', 'Avg Balance']],
                        use_container_width=True,
                        hide_index=True
                    )

            with tab2:
                st.markdown("### Unusual Spending Detection")
                st.markdown("*Transactions significantly above average (3x+ standard deviation)*")

                unusual = analyzer.detect_unusual_spending()

                if unusual:
                    st.warning(f"🚨 Found **{len(unusual)}** unusual transactions")

                    for i, txn in enumerate(unusual, 1):
                        with st.expander(f"#{i} - {txn['date']}: {format_currency(txn['amount'])} - {txn['description'][:60]}"):
                            unusual_cols = st.columns(3)
                            with unusual_cols[0]:
                                st.metric("Amount", format_currency(txn['amount']))
                            with unusual_cols[1]:
                                st.metric("Deviation", f"{txn['deviation_multiplier']:.1f}x above avg")
                            with unusual_cols[2]:
                                st.metric("Balance After", format_currency(txn['balance_after']))

                            st.info(f"**Type:** {txn['type']}")
                            st.info(f"**Description:** {txn['description']}")
                            st.info(f"**Mean Spending:** {format_currency(txn['mean_spending'])}")
                else:
                    st.success("✅ No unusual spending detected")

            with tab3:
                st.markdown("### Gambling Activity Analysis")
                st.markdown("*Detection of betting/casino transactions*")

                gambling = analyzer.analyze_gambling_activity()

                if gambling['detected']:
                    st.error(f"⚠️ Gambling activity detected: **{gambling['transaction_count']}** transactions")

                    gamble_cols = st.columns(4)
                    with gamble_cols[0]:
                        st.metric("Total Spent", format_currency(gambling['total_spent']))
                    with gamble_cols[1]:
                        st.metric("Total Won", format_currency(gambling['total_won']))
                    with gamble_cols[2]:
                        st.metric("Net Loss", format_currency(gambling['net_loss']))
                    with gamble_cols[3]:
                        st.metric("Frequency", f"{gambling['frequency']:.2f}/day")

                    st.markdown(f"**Period:** {gambling['date_range']['first']} to {gambling['date_range']['last']}")

                    # Show transactions
                    st.markdown("#### Gambling Transactions")
                    gamble_df = pd.DataFrame(gambling['transactions'])
                    gamble_df['Date'] = gamble_df['date']
                    gamble_df['Description'] = gamble_df['description']
                    gamble_df['Spent'] = gamble_df['money_out'].apply(lambda x: format_currency(x) if x > 0 else '-')
                    gamble_df['Won'] = gamble_df['money_in'].apply(lambda x: format_currency(x) if x > 0 else '-')

                    st.dataframe(
                        gamble_df[['Date', 'Description', 'Spent', 'Won']],
                        use_container_width=True,
                        hide_index=True
                    )
                else:
                    st.success("✅ No gambling activity detected")

            with tab4:
                st.markdown("### Fraud Indicators")
                st.markdown("*Patterns consistent with APP fraud or financial coercion*")

                fraud = analyzer.detect_fraud_indicators()

                # Rapid transfers
                if fraud['rapid_large_transfers']:
                    st.error(f"🚨 **{len(fraud['rapid_large_transfers'])}** days with multiple large transfers (>£1,000)")

                    for event in fraud['rapid_large_transfers']:
                        with st.expander(f"{event['date']}: {event['count']} transfers totaling {format_currency(event['total_amount'])}"):
                            for txn in event['transactions']:
                                st.write(f"• {txn['description']}: {format_currency(txn['amount'])}")
                else:
                    st.success("✅ No rapid large transfers detected")

                # Account drained
                if fraud['account_drained']:
                    st.error("🚨 **Account drain detected** (>80% balance drop)")

                    for drain in fraud['drain_events']:
                        st.warning(f"**{drain['date']}**: {drain['drop_percentage']:.1f}% drop ({format_currency(drain['amount'])}) - {drain['description']}")
                else:
                    st.success("✅ No account drain events")

                # Unusual transfer activity
                if fraud['unusual_transfer_activity']:
                    st.warning(f"⚠️ **High transfer frequency**: {fraud['transfer_rate_per_day']:.1f} transfers/day (>2.0 threshold)")
                else:
                    st.success("✅ Transfer frequency normal")

                # Coercion patterns
                st.markdown("#### Financial Coercion Indicators")
                coercion = analyzer.detect_financial_coercion_patterns()

                if coercion['sudden_pattern_changes']:
                    st.warning("⚠️ **Sudden spending pattern change detected**")
                    for change in coercion['sudden_pattern_changes']:
                        st.write(f"• {change['description']}")
                        st.write(f"  First half avg: {format_currency(change['first_half_avg'])}")
                        st.write(f"  Second half avg: {format_currency(change['second_half_avg'])}")
                        st.write(f"  Change: {change['change_percentage']:+.1f}%")

                volatility = coercion['balance_volatility']
                volatility_status = "High" if volatility > 0.5 else "Normal"
                volatility_color = "error" if volatility > 0.5 else "success"

                getattr(st, volatility_color)(f"**Balance Volatility:** {volatility:.2f} ({volatility_status})")

            with tab5:
                st.markdown("### Lifestyle Spending Analysis")
                st.markdown("*Categorized spending patterns for legal assessments*")

                lifestyle = analyzer.analyze_lifestyle_spending()

                st.info(f"✓ Categorized **{lifestyle['categorized_percentage']:.1f}%** of spending")

                # Top categories
                st.markdown("#### Spending by Category")

                categories_data = []
                for category, data in lifestyle['categories'].items():
                    categories_data.append({
                        'Category': category.title(),
                        'Total': format_currency(data['total']),
                        'Transactions': data['count'],
                        'Avg per Transaction': format_currency(data['total'] / data['count'])
                    })

                if categories_data:
                    cat_df = pd.DataFrame(categories_data)
                    st.dataframe(cat_df, use_container_width=True, hide_index=True)

                # Income analysis
                st.markdown("#### Income Sources")
                income = analyzer.analyze_income_sources()

                income_cols = st.columns(4)
                with income_cols[0]:
                    st.metric("Total Income", format_currency(income['total_income']))
                with income_cols[1]:
                    st.metric("Average Incoming", format_currency(income['average_incoming']))
                with income_cols[2]:
                    st.metric("Regular Streams", len(income['regular_income_streams']))
                with income_cols[3]:
                    st.metric("Income Transactions", income['transaction_count'])

                # Income breakdown
                income_breakdown = []
                for itype, amount in income['by_type'].items():
                    if amount > 0:
                        income_breakdown.append({
                            'Type': itype.title(),
                            'Amount': format_currency(amount),
                            'Percentage': f"{(amount / income['total_income'] * 100):.1f}%"
                        })

                if income_breakdown:
                    st.markdown("**Income by Type:**")
                    inc_df = pd.DataFrame(income_breakdown)
                    st.dataframe(inc_df, use_container_width=True, hide_index=True)

                # Regular income streams
                if income['regular_income_streams']:
                    st.markdown("**Regular Income Streams:**")
                    for stream in income['regular_income_streams']:
                        with st.expander(f"{stream['description']}: {format_currency(stream['amount'])} every {stream['frequency_days']:.0f} days"):
                            st.write(f"**Occurrences:** {stream['occurrences']}")
                            st.write(f"**Total:** {format_currency(stream['total'])}")
                            st.write(f"**Period:** {stream['first_date']} to {stream['last_date']}")

        except ValueError as e:
            st.info(f"ℹ️ Analytics unavailable: {str(e)}")
        except Exception as e:
            st.error(f"Error running analytics: {str(e)}")

        # Export section
        st.markdown("### 💾 Export Data")
        exp_cols = st.columns(2)

        with exp_cols[0]:
            # Generate Excel
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Transactions"

            # Headers
            headers = ['Date', 'Description', 'Money In', 'Money Out', 'Balance', 'Type']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Data
            for row, txn in enumerate(result.transactions, 2):
                ws.cell(row=row, column=1, value=txn.date.strftime('%d/%m/%Y') if txn.date else '')
                ws.cell(row=row, column=2, value=txn.description)
                ws.cell(row=row, column=3, value=txn.money_in if txn.money_in > 0 else None)
                ws.cell(row=row, column=4, value=txn.money_out if txn.money_out > 0 else None)
                ws.cell(row=row, column=5, value=txn.balance)
                ws.cell(row=row, column=6, value=txn.transaction_type.value if txn.transaction_type else '')

            # Column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 20

            output = BytesIO()
            wb.save(output)

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            bank_name = result.statement.bank_name or 'statement'

            st.download_button(
                "📥 Download Excel",
                data=output.getvalue(),
                file_name=f"{bank_name}_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with exp_cols[1]:
            csv_data = df.to_csv(index=False)
            st.download_button(
                "📥 Download CSV",
                data=csv_data,
                file_name=f"{bank_name}_{timestamp}.csv",
                mime="text/csv",
                use_container_width=True
            )

    # Footer
    st.markdown("---")
    st.markdown("""
        <div style='text-align: center; padding: 2rem 0; opacity: 0.6;'>
            <p>Built for <strong>Fifty Six Law</strong> | Powered by AI & Python</p>
            <p style='font-size: 0.9rem;'>🔒 All processing is local. Your data never leaves this system.</p>
        </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
