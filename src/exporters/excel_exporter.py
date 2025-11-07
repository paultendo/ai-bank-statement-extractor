"""
Excel exporter for bank statement data.

Generates Excel workbook with 3 sheets:
1. Transactions - Main transaction data
2. Statement Metadata - Account and period information
3. Extraction Log - Audit trail and confidence scores
"""
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

from ..models import ExtractionResult

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export extraction results to formatted Excel workbook."""

    # Colors
    HEADER_COLOR = "366092"  # Dark blue
    WARNING_COLOR = "FFC7CE"  # Light red
    SUCCESS_COLOR = "C6EFCE"  # Light green
    INFO_COLOR = "FFEB9C"  # Light yellow

    def __init__(self):
        """Initialize Excel exporter."""
        if openpyxl is None:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

    def export(
        self,
        result: ExtractionResult,
        output_path: Path,
        highlight_low_confidence: bool = True,
        confidence_threshold: float = 70.0
    ) -> Path:
        """
        Export extraction result to Excel.

        Args:
            result: Extraction result to export
            output_path: Path for output Excel file
            highlight_low_confidence: Whether to highlight low-confidence rows
            confidence_threshold: Confidence score below which to highlight

        Returns:
            Path to created Excel file
        """
        logger.info(f"Exporting to Excel: {output_path}")

        # Create workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create sheets
        self._create_transactions_sheet(
            wb,
            result,
            highlight_low_confidence,
            confidence_threshold
        )
        self._create_metadata_sheet(wb, result)
        self._create_audit_log_sheet(wb, result)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel export complete: {output_path}")

        return output_path

    def _get_currency_format(self, result: ExtractionResult) -> str:
        """
        Get Excel currency format string based on statement currency.

        Args:
            result: Extraction result with statement metadata

        Returns:
            Excel number format string (e.g., '£#,##0.00', 'R$#,##0.00')
        """
        # Default to GBP if no statement metadata
        if not result.statement or not result.statement.currency:
            return '£#,##0.00'

        currency = result.statement.currency.upper()

        # Map currency codes to Excel format strings
        currency_formats = {
            'GBP': '£#,##0.00',
            'EUR': '€#,##0.00',
            'USD': '$#,##0.00',
            'BRL': 'R$#,##0.00',  # Brazilian Real
            'CAD': 'CA$#,##0.00',
            'AUD': 'A$#,##0.00',
            'JPY': '¥#,##0',  # No decimal places for Yen
            'CHF': 'CHF#,##0.00',
            'INR': '₹#,##0.00',
        }

        return currency_formats.get(currency, f'{currency} #,##0.00')

    def _create_transactions_sheet(
        self,
        wb: openpyxl.Workbook,
        result: ExtractionResult,
        highlight_low_confidence: bool,
        confidence_threshold: float
    ) -> None:
        """Create transactions sheet with formatted data."""
        ws = wb.create_sheet("Transactions", 0)

        # Check if any transactions have translations
        has_translations = any(
            txn.description_translated for txn in result.transactions
        )

        # Determine currency format from statement metadata
        currency_format = self._get_currency_format(result)

        # Headers (conditionally include translation column)
        headers = ["Date", "Description"]
        if has_translations:
            headers.append("Description (English)")
        headers.extend(["Money In", "Money Out", "Balance", "Type", "Confidence %"])

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.HEADER_COLOR, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write transactions
        for row, txn in enumerate(result.transactions, 2):
            col_idx = 1
            ws.cell(row=row, column=col_idx, value=txn.date.strftime("%Y-%m-%d"))
            col_idx += 1

            ws.cell(row=row, column=col_idx, value=txn.description)
            col_idx += 1

            if has_translations:
                ws.cell(row=row, column=col_idx, value=txn.description_translated or "")
                col_idx += 1

            ws.cell(row=row, column=col_idx, value=txn.money_in if txn.money_in > 0 else "")
            money_in_col = col_idx
            col_idx += 1

            ws.cell(row=row, column=col_idx, value=txn.money_out if txn.money_out > 0 else "")
            money_out_col = col_idx
            col_idx += 1

            ws.cell(row=row, column=col_idx, value=txn.balance)
            balance_col = col_idx
            col_idx += 1

            ws.cell(row=row, column=col_idx, value=txn.transaction_type.value if txn.transaction_type else "")
            col_idx += 1

            ws.cell(row=row, column=col_idx, value=round(txn.confidence, 1))

            # Format numbers as currency
            for col in [money_in_col, money_out_col, balance_col]:
                ws.cell(row=row, column=col).number_format = currency_format

            # Highlight low confidence rows
            if highlight_low_confidence and txn.confidence < confidence_threshold:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color=self.WARNING_COLOR,
                        fill_type="solid"
                    )

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Description columns wider
        ws.column_dimensions['B'].width = 50
        if has_translations:
            ws.column_dimensions['C'].width = 50

        # Add totals row
        total_row = len(result.transactions) + 3
        ws.cell(row=total_row, column=1, value="TOTALS")
        ws.cell(row=total_row, column=1).font = Font(bold=True)

        total_in = sum(txn.money_in for txn in result.transactions)
        total_out = sum(txn.money_out for txn in result.transactions)

        # Totals in correct columns (account for optional translation column)
        ws.cell(row=total_row, column=money_in_col, value=total_in)
        ws.cell(row=total_row, column=money_out_col, value=total_out)

        for col in [money_in_col, money_out_col]:
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            cell.number_format = currency_format
            cell.fill = PatternFill(start_color=self.INFO_COLOR, fill_type="solid")

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_metadata_sheet(
        self,
        wb: openpyxl.Workbook,
        result: ExtractionResult
    ) -> None:
        """Create statement metadata sheet."""
        ws = wb.create_sheet("Statement Metadata", 1)

        if not result.statement:
            ws.cell(row=1, column=1, value="No statement metadata available")
            return

        stmt = result.statement

        # Get currency symbol for display
        currency_symbols = {
            'GBP': '£',
            'EUR': '€',
            'USD': '$',
            'BRL': 'R$',
            'CAD': 'CA$',
            'AUD': 'A$',
            'JPY': '¥',
            'CHF': 'CHF',
            'INR': '₹',
        }
        currency_symbol = currency_symbols.get(stmt.currency.upper(), stmt.currency)

        # Metadata rows
        metadata = [
            ("Bank Name", stmt.bank_name),
            ("Account Number", stmt.account_number),
            ("Account Holder", stmt.account_holder or "N/A"),
            ("Sort Code", stmt.sort_code or "N/A"),
            ("Currency", stmt.currency),
            ("", ""),
            ("Statement Period", ""),
            ("  Start Date", stmt.statement_start_date.strftime("%Y-%m-%d")),
            ("  End Date", stmt.statement_end_date.strftime("%Y-%m-%d")),
            ("", ""),
            ("Balances", ""),
            ("  Opening Balance", f"{currency_symbol}{stmt.opening_balance:,.2f}"),
            ("  Closing Balance", f"{currency_symbol}{stmt.closing_balance:,.2f}"),
            ("", ""),
            ("Transaction Count", len(result.transactions)),
            ("Extraction Method", result.extraction_method),
            ("Confidence Score", f"{result.confidence_score:.1f}%"),
            ("Balance Reconciled", "✓ Yes" if result.balance_reconciled else "✗ No"),
            ("", ""),
            ("Processing Time", f"{result.processing_time:.2f} seconds"),
            ("Extracted At", result.extracted_at.strftime("%Y-%m-%d %H:%M:%S")),
        ]

        # Write metadata
        for row, (label, value) in enumerate(metadata, 1):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)

            if label and not label.startswith("  "):
                ws.cell(row=row, column=1).font = Font(bold=True)

        # Color the reconciliation row
        recon_row = 18
        if result.balance_reconciled:
            ws.cell(row=recon_row, column=2).fill = PatternFill(
                start_color=self.SUCCESS_COLOR,
                fill_type="solid"
            )
        else:
            ws.cell(row=recon_row, column=2).fill = PatternFill(
                start_color=self.WARNING_COLOR,
                fill_type="solid"
            )

        # Auto-size columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30

    def _create_audit_log_sheet(
        self,
        wb: openpyxl.Workbook,
        result: ExtractionResult
    ) -> None:
        """Create audit log sheet."""
        ws = wb.create_sheet("Extraction Log", 2)

        # Headers
        ws.cell(row=1, column=1, value="Extraction Audit Log")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        row = 3

        # Overall status
        ws.cell(row=row, column=1, value="Status:")
        ws.cell(row=row, column=2, value="SUCCESS" if result.success else "FAILED")
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).fill = PatternFill(
            start_color=self.SUCCESS_COLOR if result.success else self.WARNING_COLOR,
            fill_type="solid"
        )
        row += 2

        # Extraction method
        ws.cell(row=row, column=1, value="Extraction Method:")
        ws.cell(row=row, column=2, value=result.extraction_method)
        row += 1

        # Confidence
        ws.cell(row=row, column=1, value="Overall Confidence:")
        ws.cell(row=row, column=2, value=f"{result.confidence_score:.1f}%")
        row += 2

        # Warnings
        if result.warnings:
            ws.cell(row=row, column=1, value="Warnings:")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            for warning in result.warnings:
                ws.cell(row=row, column=2, value=f"⚠ {warning}")
                ws.cell(row=row, column=2).fill = PatternFill(
                    start_color=self.INFO_COLOR,
                    fill_type="solid"
                )
                row += 1

            row += 1

        # Error message
        if result.error_message:
            ws.cell(row=row, column=1, value="Error:")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=result.error_message)
            ws.cell(row=row, column=2).fill = PatternFill(
                start_color=self.WARNING_COLOR,
                fill_type="solid"
            )
            row += 2

        # Low confidence transactions
        low_conf = result.low_confidence_transactions
        if low_conf:
            ws.cell(row=row, column=1, value=f"Low Confidence Transactions ({len(low_conf)}):")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            ws.cell(row=row, column=1, value="Date")
            ws.cell(row=row, column=2, value="Description")
            ws.cell(row=row, column=3, value="Confidence %")

            for cell in [ws.cell(row=row, column=i) for i in [1, 2, 3]]:
                cell.font = Font(bold=True)

            row += 1

            for txn in low_conf[:20]:  # Limit to 20
                ws.cell(row=row, column=1, value=txn.date.strftime("%Y-%m-%d"))
                ws.cell(row=row, column=2, value=txn.description[:50])
                ws.cell(row=row, column=3, value=f"{txn.confidence:.1f}%")
                row += 1

        # Auto-size columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15


def generate_output_filename(
    bank_name: str,
    statement_date: datetime,
    output_dir: Optional[Path] = None
) -> Path:
    """
    Generate standardized output filename.

    Format: {bank}_{YYYY-MM-DD}_{timestamp}.xlsx

    Args:
        bank_name: Name of bank
        statement_date: Statement date
        output_dir: Output directory (default: ./output)

    Returns:
        Path for output file
    """
    from ..config.settings import OUTPUT_DIR

    if output_dir is None:
        output_dir = OUTPUT_DIR

    timestamp = datetime.now().strftime("%H%M%S")
    date_str = statement_date.strftime("%Y-%m-%d")
    filename = f"{bank_name.lower()}_{date_str}_{timestamp}.xlsx"

    return output_dir / filename
